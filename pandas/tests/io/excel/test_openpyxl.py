import pytest

import pandas.util._test_decorators as td

from pandas import DataFrame
from pandas.util.testing import ensure_clean

from pandas.io.excel import ExcelWriter, _OpenpyxlWriter


@td.skip_if_no('openpyxl')
@pytest.mark.parametrize("ext", ['.xlsx'])
class TestOpenpyxlTests:

    def test_to_excel_styleconverter(self, ext):
        from openpyxl import styles

        hstyle = {
            "font": {
                "color": '00FF0000',
                "bold": True,
            },
            "borders": {
                "top": "thin",
                "right": "thin",
                "bottom": "thin",
                "left": "thin",
            },
            "alignment": {
                "horizontal": "center",
                "vertical": "top",
            },
            "fill": {
                "patternType": 'solid',
                'fgColor': {
                    'rgb': '006666FF',
                    'tint': 0.3,
                },
            },
            "number_format": {
                "format_code": "0.00"
            },
            "protection": {
                "locked": True,
                "hidden": False,
            },
        }

        font_color = styles.Color('00FF0000')
        font = styles.Font(bold=True, color=font_color)
        side = styles.Side(style=styles.borders.BORDER_THIN)
        border = styles.Border(top=side, right=side, bottom=side, left=side)
        alignment = styles.Alignment(horizontal='center', vertical='top')
        fill_color = styles.Color(rgb='006666FF', tint=0.3)
        fill = styles.PatternFill(patternType='solid', fgColor=fill_color)

        number_format = '0.00'

        protection = styles.Protection(locked=True, hidden=False)

        kw = _OpenpyxlWriter._convert_to_style_kwargs(hstyle)
        assert kw['font'] == font
        assert kw['border'] == border
        assert kw['alignment'] == alignment
        assert kw['fill'] == fill
        assert kw['number_format'] == number_format
        assert kw['protection'] == protection

    def test_write_cells_merge_styled(self, ext):
        from pandas.io.formats.excel import ExcelCell

        sheet_name = 'merge_styled'

        sty_b1 = {'font': {'color': '00FF0000'}}
        sty_a2 = {'font': {'color': '0000FF00'}}

        initial_cells = [
            ExcelCell(col=1, row=0, val=42, style=sty_b1),
            ExcelCell(col=0, row=1, val=99, style=sty_a2),
        ]

        sty_merged = {'font': {'color': '000000FF', 'bold': True}}
        sty_kwargs = _OpenpyxlWriter._convert_to_style_kwargs(sty_merged)
        openpyxl_sty_merged = sty_kwargs['font']
        merge_cells = [
            ExcelCell(col=0, row=0, val='pandas',
                      mergestart=1, mergeend=1, style=sty_merged),
        ]

        with ensure_clean(ext) as path:
            writer = _OpenpyxlWriter(path)
            writer.write_cells(initial_cells, sheet_name=sheet_name)
            writer.write_cells(merge_cells, sheet_name=sheet_name)

            wks = writer.sheets[sheet_name]
            xcell_b1 = wks['B1']
            xcell_a2 = wks['A2']
            assert xcell_b1.font == openpyxl_sty_merged
            assert xcell_a2.font == openpyxl_sty_merged

    @pytest.mark.parametrize("mode,expected", [
        ('w', ['baz']), ('a', ['foo', 'bar', 'baz'])])
    def test_write_append_mode(self, ext, mode, expected):
        import openpyxl
        df = DataFrame([1], columns=['baz'])

        with ensure_clean(ext) as f:
            wb = openpyxl.Workbook()
            wb.worksheets[0].title = 'foo'
            wb.worksheets[0]['A1'].value = 'foo'
            wb.create_sheet('bar')
            wb.worksheets[1]['A1'].value = 'bar'
            wb.save(f)

            writer = ExcelWriter(f, engine='openpyxl', mode=mode)
            df.to_excel(writer, sheet_name='baz', index=False)
            writer.save()

            wb2 = openpyxl.load_workbook(f)
            result = [sheet.title for sheet in wb2.worksheets]
            assert result == expected

            for index, cell_value in enumerate(expected):
                assert wb2.worksheets[index]['A1'].value == cell_value
